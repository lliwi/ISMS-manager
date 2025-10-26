"""
Blueprint para gestión de incidentes de seguridad
ISO 27001:2023 - Controles 5.24, 5.25, 5.26, 5.27, 5.28
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from models import (
    Incident, IncidentCategory, IncidentSeverity, IncidentPriority,
    IncidentStatus, IncidentSource, DetectionMethod,
    IncidentTimeline, ActionType, IncidentAction, ActionStatus,
    IncidentEvidence, EvidenceType, IncidentNotification, NotificationType,
    IncidentAsset, Asset, User, db
)
from datetime import datetime, timedelta
from sqlalchemy import or_, and_, func
import hashlib
import os
from werkzeug.utils import secure_filename

incidents_bp = Blueprint('incidents', __name__, url_prefix='/incidents')

# Configuración de uploads
UPLOAD_FOLDER = 'uploads/incidents'
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'log', 'pcap', 'zip'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def calculate_file_hash(file_path):
    """Calcula SHA-256 hash de un archivo para integridad"""
    sha256_hash = hashlib.sha256()
    with open(file_path, "rb") as f:
        for byte_block in iter(lambda: f.read(4096), b""):
            sha256_hash.update(byte_block)
    return sha256_hash.hexdigest()


# ============================================================================
# VISTAS PRINCIPALES
# ============================================================================

@incidents_bp.route('/')
@login_required
def index():
    """Dashboard de incidentes"""
    # Filtros
    status_filter = request.args.get('status')
    category_filter = request.args.get('category')
    severity_filter = request.args.get('severity')
    search = request.args.get('search', '')

    # Query base
    query = Incident.query

    # Aplicar filtros
    if status_filter and status_filter != 'all':
        query = query.filter(Incident.status == IncidentStatus[status_filter])

    if category_filter and category_filter != 'all':
        query = query.filter(Incident.category == IncidentCategory[category_filter])

    if severity_filter and severity_filter != 'all':
        query = query.filter(Incident.severity == IncidentSeverity[severity_filter])

    if search:
        query = query.filter(
            or_(
                Incident.incident_number.ilike(f'%{search}%'),
                Incident.title.ilike(f'%{search}%'),
                Incident.description.ilike(f'%{search}%')
            )
        )

    # Ordenar y paginar
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    incidents_pagination = query.order_by(Incident.reported_date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    # Estadísticas rápidas
    stats = {
        'total': Incident.query.count(),
        'open': Incident.query.filter(Incident.status.in_([
            IncidentStatus.NEW, IncidentStatus.EVALUATING,
            IncidentStatus.CONFIRMED, IncidentStatus.IN_PROGRESS
        ])).count(),
        'critical': Incident.query.filter(Incident.severity == IncidentSeverity.CRITICAL).count(),
        'this_month': Incident.query.filter(
            Incident.reported_date >= datetime.utcnow().replace(day=1)
        ).count()
    }

    return render_template('incidents/index.html',
                          incidents=incidents_pagination.items,
                          pagination=incidents_pagination,
                          stats=stats,
                          IncidentStatus=IncidentStatus,
                          IncidentCategory=IncidentCategory,
                          IncidentSeverity=IncidentSeverity)


@incidents_bp.route('/new', methods=['GET', 'POST'])
@login_required
def create():
    """Crear/reportar nuevo incidente"""
    if request.method == 'POST':
        try:
            # Generar número de incidente
            incident_number = Incident.generate_incident_number()

            # Crear incidente
            incident = Incident(
                incident_number=incident_number,
                title=request.form['title'],
                description=request.form['description'],
                category=IncidentCategory[request.form['category']],
                severity=IncidentSeverity[request.form['severity']],
                priority=IncidentPriority[request.form['priority']],
                status=IncidentStatus.NEW,
                discovery_date=datetime.strptime(request.form['discovery_date'], '%Y-%m-%d'),
                reported_date=datetime.utcnow(),
                reported_by_id=current_user.id,
                source=IncidentSource[request.form.get('source', 'UNKNOWN')],
                detection_method=DetectionMethod[request.form['detection_method']],
                detection_details=request.form.get('detection_details'),
                impact_confidentiality=request.form.get('impact_confidentiality') == 'on',
                impact_integrity=request.form.get('impact_integrity') == 'on',
                impact_availability=request.form.get('impact_availability') == 'on',
                created_by_id=current_user.id
            )

            # Asignar automáticamente si se especifica
            assigned_to_id = request.form.get('assigned_to_id')
            if assigned_to_id:
                incident.assigned_to_id = int(assigned_to_id)

            db.session.add(incident)
            db.session.flush()  # Para obtener el ID

            # Agregar activos afectados
            affected_asset_ids = request.form.getlist('affected_assets[]')
            for asset_id in affected_asset_ids:
                if asset_id:
                    incident_asset = IncidentAsset(
                        incident_id=incident.id,
                        asset_id=int(asset_id)
                    )
                    db.session.add(incident_asset)

            # Registrar evento de creación en timeline
            timeline_event = IncidentTimeline(
                incident_id=incident.id,
                timestamp=datetime.utcnow(),
                action_type=ActionType.CREATED,
                description=f"Incidente creado por {current_user.name}",
                user_id=current_user.id
            )
            db.session.add(timeline_event)

            db.session.commit()

            flash(f'Incidente {incident.incident_number} reportado correctamente', 'success')
            return redirect(url_for('incidents.view', id=incident.id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear incidente: {str(e)}', 'error')

    # GET: Mostrar formulario
    users = User.query.filter_by(is_active=True).all()
    assets = Asset.query.all()

    return render_template('incidents/create.html',
                          users=users,
                          assets=assets,
                          now=datetime.utcnow(),
                          IncidentCategory=IncidentCategory,
                          IncidentSeverity=IncidentSeverity,
                          IncidentPriority=IncidentPriority,
                          IncidentSource=IncidentSource,
                          DetectionMethod=DetectionMethod)


@incidents_bp.route('/<int:id>')
@login_required
def view(id):
    """Ver detalle de incidente"""
    incident = Incident.query.get_or_404(id)

    # Obtener métricas
    metrics = {
        'response_time': incident.calculate_response_time(),
        'resolution_time': incident.calculate_resolution_time()
    }

    # Obtener usuarios para asignación
    users = User.query.filter_by(is_active=True).all()

    return render_template('incidents/view.html',
                          incident=incident,
                          metrics=metrics,
                          users=users,
                          IncidentStatus=IncidentStatus)


@incidents_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    """Editar incidente"""
    incident = Incident.query.get_or_404(id)

    # Solo el asignado, reportador o admin pueden editar
    if not (current_user.id == incident.assigned_to_id or
            current_user.id == incident.reported_by_id or
            current_user.has_role('admin')):
        flash('No tienes permisos para editar este incidente', 'error')
        return redirect(url_for('incidents.view', id=id))

    if request.method == 'POST':
        try:
            old_status = incident.status

            # Actualizar campos
            incident.title = request.form['title']
            incident.description = request.form['description']
            incident.category = IncidentCategory[request.form['category']]
            incident.severity = IncidentSeverity[request.form['severity']]
            incident.priority = IncidentPriority[request.form['priority']]
            incident.source = IncidentSource[request.form.get('source', 'UNKNOWN')]
            incident.detection_method = DetectionMethod[request.form['detection_method']]
            incident.detection_details = request.form.get('detection_details')
            incident.impact_confidentiality = request.form.get('impact_confidentiality') == 'on'
            incident.impact_integrity = request.form.get('impact_integrity') == 'on'
            incident.impact_availability = request.form.get('impact_availability') == 'on'
            incident.root_cause = request.form.get('root_cause')
            incident.contributing_factors = request.form.get('contributing_factors')
            incident.resolution = request.form.get('resolution')
            incident.lessons_learned = request.form.get('lessons_learned')
            incident.updated_by_id = current_user.id

            # Registrar cambio en timeline
            timeline_event = IncidentTimeline(
                incident_id=incident.id,
                timestamp=datetime.utcnow(),
                action_type=ActionType.COMMENT,
                description=f"Incidente actualizado por {current_user.name}",
                user_id=current_user.id
            )
            db.session.add(timeline_event)

            db.session.commit()
            flash('Incidente actualizado correctamente', 'success')
            return redirect(url_for('incidents.view', id=id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar incidente: {str(e)}', 'error')

    users = User.query.filter_by(is_active=True).all()
    assets = Asset.query.all()

    return render_template('incidents/edit.html',
                          incident=incident,
                          users=users,
                          assets=assets,
                          now=datetime.utcnow(),
                          IncidentCategory=IncidentCategory,
                          IncidentSeverity=IncidentSeverity,
                          IncidentPriority=IncidentPriority,
                          IncidentSource=IncidentSource,
                          DetectionMethod=DetectionMethod)


# ============================================================================
# GESTIÓN DE ESTADO
# ============================================================================

@incidents_bp.route('/<int:id>/status', methods=['POST'])
@login_required
def change_status(id):
    """Cambiar estado del incidente"""
    incident = Incident.query.get_or_404(id)

    try:
        new_status = IncidentStatus[request.form['status']]
        old_status = incident.status
        comment = request.form.get('comment', '')

        incident.status = new_status
        incident.updated_by_id = current_user.id

        # Actualizar fechas según el estado
        if new_status == IncidentStatus.CONTAINED and not incident.containment_date:
            incident.containment_date = datetime.utcnow()
        elif new_status == IncidentStatus.RESOLVED and not incident.resolution_date:
            incident.resolution_date = datetime.utcnow()
        elif new_status == IncidentStatus.CLOSED and not incident.closure_date:
            incident.closure_date = datetime.utcnow()

        # Registrar en timeline
        description = f"Estado cambiado de {old_status.value} a {new_status.value}"
        if comment:
            description += f". Comentario: {comment}"

        timeline_event = IncidentTimeline(
            incident_id=incident.id,
            timestamp=datetime.utcnow(),
            action_type=ActionType.STATUS_CHANGE,
            description=description,
            details=comment,
            user_id=current_user.id
        )
        db.session.add(timeline_event)

        db.session.commit()
        flash(f'Estado actualizado a {new_status.value}', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al cambiar estado: {str(e)}', 'error')

    return redirect(url_for('incidents.view', id=id))


@incidents_bp.route('/<int:id>/assign', methods=['POST'])
@login_required
def assign(id):
    """Asignar incidente a un usuario"""
    incident = Incident.query.get_or_404(id)

    try:
        user_id = request.form.get('user_id')
        if user_id:
            user = User.query.get(int(user_id))
            old_assigned = incident.assigned_to

            incident.assigned_to_id = user.id
            incident.updated_by_id = current_user.id

            # Registrar en timeline
            description = f"Incidente asignado a {user.name}"
            if old_assigned:
                description = f"Incidente reasignado de {old_assigned.name} a {user.name}"

            timeline_event = IncidentTimeline(
                incident_id=incident.id,
                timestamp=datetime.utcnow(),
                action_type=ActionType.ASSIGNED,
                description=description,
                user_id=current_user.id
            )
            db.session.add(timeline_event)

            db.session.commit()
            flash(f'Incidente asignado a {user.name}', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al asignar incidente: {str(e)}', 'error')

    return redirect(url_for('incidents.view', id=id))


# ============================================================================
# TIMELINE Y COMENTARIOS
# ============================================================================

@incidents_bp.route('/<int:id>/comment', methods=['POST'])
@login_required
def add_comment(id):
    """Agregar comentario al incidente"""
    incident = Incident.query.get_or_404(id)

    try:
        comment = request.form.get('comment')
        if comment:
            timeline_event = IncidentTimeline(
                incident_id=incident.id,
                timestamp=datetime.utcnow(),
                action_type=ActionType.COMMENT,
                description=comment,
                user_id=current_user.id
            )
            db.session.add(timeline_event)
            db.session.commit()
            flash('Comentario agregado', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar comentario: {str(e)}', 'error')

    return redirect(url_for('incidents.view', id=id))


# ============================================================================
# ACCIONES CORRECTIVAS
# ============================================================================

@incidents_bp.route('/<int:id>/actions', methods=['GET', 'POST'])
@login_required
def actions(id):
    """Gestionar acciones correctivas"""
    incident = Incident.query.get_or_404(id)

    if request.method == 'POST':
        try:
            action = IncidentAction(
                incident_id=incident.id,
                action_type=request.form.get('action_type', 'Correctiva'),
                description=request.form['description'],
                responsible_id=int(request.form['responsible_id']) if request.form.get('responsible_id') else None,
                due_date=datetime.strptime(request.form['due_date'], '%Y-%m-%d').date() if request.form.get('due_date') else None,
                status=ActionStatus.PENDING
            )

            db.session.add(action)

            # Registrar en timeline
            timeline_event = IncidentTimeline(
                incident_id=incident.id,
                timestamp=datetime.utcnow(),
                action_type=ActionType.ACTION_ADDED,
                description=f"Nueva acción agregada: {action.description[:50]}...",
                user_id=current_user.id
            )
            db.session.add(timeline_event)

            db.session.commit()
            flash('Acción creada correctamente', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear acción: {str(e)}', 'error')

        return redirect(url_for('incidents.view', id=id))

    users = User.query.filter_by(is_active=True).all()
    actions = incident.actions.all() if hasattr(incident, 'actions') else []

    # Estadísticas de acciones
    stats = {
        'total': len(actions),
        'pending': len([a for a in actions if a.status == ActionStatus.PENDING]),
        'in_progress': len([a for a in actions if a.status == ActionStatus.IN_PROGRESS]),
        'completed': len([a for a in actions if a.status == ActionStatus.COMPLETED]),
        'overdue': len([a for a in actions if hasattr(a, 'is_overdue') and a.is_overdue])
    }

    return render_template('incidents/actions.html',
                          incident=incident,
                          actions=actions,
                          users=users,
                          stats=stats,
                          now=datetime.utcnow(),
                          ActionStatus=ActionStatus)


@incidents_bp.route('/actions/<int:action_id>/complete', methods=['POST'])
@login_required
def complete_action(action_id):
    """Marcar acción como completada"""
    action = IncidentAction.query.get_or_404(action_id)

    try:
        action.status = ActionStatus.COMPLETED
        action.completion_date = datetime.utcnow().date()
        action.notes = request.form.get('notes')

        # Registrar en timeline
        timeline_event = IncidentTimeline(
            incident_id=action.incident_id,
            timestamp=datetime.utcnow(),
            action_type=ActionType.COMMENT,
            description=f"Acción completada: {action.description[:50]}...",
            user_id=current_user.id
        )
        db.session.add(timeline_event)

        db.session.commit()
        flash('Acción marcada como completada', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al completar acción: {str(e)}', 'error')

    return redirect(url_for('incidents.view', id=action.incident_id))


@incidents_bp.route('/<int:id>/actions/<int:action_id>/status', methods=['POST'])
@login_required
def update_action_status(id, action_id):
    """Actualizar estado de una acción"""
    incident = Incident.query.get_or_404(id)
    action = IncidentAction.query.get_or_404(action_id)

    if action.incident_id != incident.id:
        flash('Acción no pertenece a este incidente', 'error')
        return redirect(url_for('incidents.view', id=id))

    try:
        new_status = ActionStatus[request.form['status']]
        old_status = action.status

        action.status = new_status

        if new_status == ActionStatus.COMPLETED:
            action.completion_date = datetime.utcnow().date()

        # Registrar en timeline
        timeline_event = IncidentTimeline(
            incident_id=incident.id,
            timestamp=datetime.utcnow(),
            action_type=ActionType.COMMENT,
            description=f"Acción '{action.description[:30]}...' cambió de {old_status.value} a {new_status.value}",
            user_id=current_user.id
        )
        db.session.add(timeline_event)

        db.session.commit()
        flash(f'Estado de acción actualizado a {new_status.value}', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar estado: {str(e)}', 'error')

    return redirect(url_for('incidents.actions', id=id))


@incidents_bp.route('/<int:id>/actions/<int:action_id>/progress', methods=['POST'])
@login_required
def update_action_progress(id, action_id):
    """Actualizar progreso de una acción"""
    incident = Incident.query.get_or_404(id)
    action = IncidentAction.query.get_or_404(action_id)

    if action.incident_id != incident.id:
        flash('Acción no pertenece a este incidente', 'error')
        return redirect(url_for('incidents.view', id=id))

    try:
        progress = int(request.form.get('progress', 0))
        action.progress = progress
        action.notes = request.form.get('notes', action.notes)

        # Si llega al 100%, marcar como completada
        if progress == 100 and action.status != ActionStatus.COMPLETED:
            action.status = ActionStatus.COMPLETED
            action.completion_date = datetime.utcnow().date()

        db.session.commit()
        flash(f'Progreso actualizado a {progress}%', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar progreso: {str(e)}', 'error')

    return redirect(url_for('incidents.actions', id=id))


@incidents_bp.route('/<int:id>/actions/<int:action_id>/notes', methods=['POST'])
@login_required
def update_action_notes(id, action_id):
    """Actualizar notas de una acción"""
    incident = Incident.query.get_or_404(id)
    action = IncidentAction.query.get_or_404(action_id)

    if action.incident_id != incident.id:
        flash('Acción no pertenece a este incidente', 'error')
        return redirect(url_for('incidents.view', id=id))

    try:
        action.notes = request.form.get('notes', '')
        db.session.commit()
        flash('Notas actualizadas', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar notas: {str(e)}', 'error')

    return redirect(url_for('incidents.actions', id=id))


# ============================================================================
# EVIDENCIAS
# ============================================================================

@incidents_bp.route('/<int:id>/evidences', methods=['GET', 'POST'])
@login_required
def evidences(id):
    """Gestionar evidencias del incidente"""
    incident = Incident.query.get_or_404(id)

    if request.method == 'POST':
        try:
            # Procesar archivo si existe
            file_path = None
            file_hash = None
            file_size = None
            file_name = None

            if 'evidence_file' in request.files:
                file = request.files['evidence_file']
                if file and file.filename and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    # Crear directorio si no existe
                    upload_dir = os.path.join(UPLOAD_FOLDER, str(incident.id))
                    os.makedirs(upload_dir, exist_ok=True)

                    # Guardar archivo con timestamp
                    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
                    file_name = f"{timestamp}_{filename}"
                    file_path = os.path.join(upload_dir, file_name)
                    file.save(file_path)

                    # Calcular hash
                    file_hash = calculate_file_hash(file_path)
                    file_size = os.path.getsize(file_path)

            # Crear evidencia
            evidence = IncidentEvidence(
                incident_id=incident.id,
                evidence_type=EvidenceType[request.form['evidence_type']],
                description=request.form['description'],
                file_path=file_path,
                file_name=file_name,
                file_size=file_size,
                hash_value=file_hash,
                collected_by_id=current_user.id,
                collection_date=datetime.utcnow(),
                chain_of_custody=f"Recopilado por {current_user.name} el {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}",
                notes=request.form.get('notes')
            )

            db.session.add(evidence)

            # Registrar en timeline
            timeline_event = IncidentTimeline(
                incident_id=incident.id,
                timestamp=datetime.utcnow(),
                action_type=ActionType.EVIDENCE_ADDED,
                description=f"Nueva evidencia agregada: {evidence.evidence_type.value}",
                user_id=current_user.id
            )
            db.session.add(timeline_event)

            db.session.commit()
            flash('Evidencia agregada correctamente', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'Error al agregar evidencia: {str(e)}', 'error')

        return redirect(url_for('incidents.view', id=id))

    users = User.query.filter_by(is_active=True).all()
    return render_template('incidents/evidences.html',
                          incident=incident,
                          evidences=incident.evidences,
                          users=users,
                          EvidenceType=EvidenceType)


@incidents_bp.route('/<int:id>/evidences/<int:evidence_id>/download')
@login_required
def download_evidence(id, evidence_id):
    """Descargar archivo de evidencia"""
    from flask import send_file

    incident = Incident.query.get_or_404(id)
    evidence = IncidentEvidence.query.get_or_404(evidence_id)

    if evidence.incident_id != incident.id:
        flash('Evidencia no pertenece a este incidente', 'error')
        return redirect(url_for('incidents.view', id=id))

    if not evidence.file_path or not os.path.exists(evidence.file_path):
        flash('Archivo de evidencia no encontrado', 'error')
        return redirect(url_for('incidents.evidences', id=id))

    try:
        # Registrar acceso en timeline
        timeline_event = IncidentTimeline(
            incident_id=incident.id,
            timestamp=datetime.utcnow(),
            action_type=ActionType.COMMENT,
            description=f"Evidencia '{evidence.file_name}' descargada por {current_user.name}",
            user_id=current_user.id
        )
        db.session.add(timeline_event)
        db.session.commit()

        return send_file(
            evidence.file_path,
            as_attachment=True,
            download_name=evidence.file_name
        )

    except Exception as e:
        flash(f'Error al descargar evidencia: {str(e)}', 'error')
        return redirect(url_for('incidents.evidences', id=id))


@incidents_bp.route('/<int:id>/evidences/<int:evidence_id>/delete', methods=['POST'])
@login_required
def delete_evidence(id, evidence_id):
    """Eliminar evidencia"""
    incident = Incident.query.get_or_404(id)
    evidence = IncidentEvidence.query.get_or_404(evidence_id)

    if evidence.incident_id != incident.id:
        flash('Evidencia no pertenece a este incidente', 'error')
        return redirect(url_for('incidents.view', id=id))

    # Solo admin o security manager pueden eliminar evidencias
    if not current_user.has_role('admin') and not current_user.has_role('ciso'):
        flash('No tiene permisos para eliminar evidencias', 'error')
        return redirect(url_for('incidents.evidences', id=id))

    try:
        # Eliminar archivo físico si existe
        if evidence.file_path and os.path.exists(evidence.file_path):
            os.remove(evidence.file_path)

        file_name = evidence.file_name

        # Registrar eliminación en timeline
        timeline_event = IncidentTimeline(
            incident_id=incident.id,
            timestamp=datetime.utcnow(),
            action_type=ActionType.COMMENT,
            description=f"Evidencia '{file_name}' eliminada por {current_user.name}",
            user_id=current_user.id
        )
        db.session.add(timeline_event)

        # Eliminar registro de base de datos
        db.session.delete(evidence)
        db.session.commit()

        flash('Evidencia eliminada correctamente', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar evidencia: {str(e)}', 'error')

    return redirect(url_for('incidents.evidences', id=id))


# ============================================================================
# REPORTES Y MÉTRICAS
# ============================================================================

@incidents_bp.route('/reports')
@login_required
def reports():
    """Dashboard de reportes y métricas"""
    import json
    from collections import defaultdict

    # Filtros de periodo
    period = request.args.get('period', 30, type=int)
    category = request.args.get('category', 'all')

    # Calcular rango de fechas
    end_date = datetime.utcnow()
    if period == 'custom':
        start_date = datetime.strptime(request.args.get('start_date'), '%Y-%m-%d')
        end_date = datetime.strptime(request.args.get('end_date'), '%Y-%m-%d')
    else:
        start_date = end_date - timedelta(days=period)

    # Query base
    query = Incident.query.filter(Incident.reported_date >= start_date)
    if category != 'all':
        query = query.filter(Incident.category == IncidentCategory[category])

    incidents = query.all()

    # Métricas principales
    total = len(incidents)
    critical = len([i for i in incidents if i.severity == IncidentSeverity.CRITICAL])
    resolved = len([i for i in incidents if i.status == IncidentStatus.RESOLVED])
    resolution_rate = (resolved / total * 100) if total > 0 else 0

    # Tiempos promedio
    response_times = [i.calculate_response_time() for i in incidents if i.calculate_response_time()]
    containment_times = [(i.containment_date - i.discovery_date).total_seconds() / 3600
                         for i in incidents if i.containment_date and i.discovery_date]
    resolution_times = [i.calculate_resolution_time() for i in incidents if i.calculate_resolution_time()]
    closure_times = [(i.closure_date - i.discovery_date).total_seconds() / 3600
                     for i in incidents if i.closure_date and i.discovery_date]

    avg_response_time = sum(response_times) / len(response_times) if response_times else 0
    avg_containment_time = sum(containment_times) / len(containment_times) if containment_times else 0
    avg_resolution_time = sum(resolution_times) / len(resolution_times) if resolution_times else 0
    avg_closure_time = sum(closure_times) / len(closure_times) if closure_times else 0

    # Tendencias
    prev_start = start_date - timedelta(days=period)
    prev_incidents = Incident.query.filter(
        Incident.reported_date >= prev_start,
        Incident.reported_date < start_date
    ).count()
    trend_total = ((total - prev_incidents) / prev_incidents * 100) if prev_incidents > 0 else 0

    # Brechas de datos RGPD
    data_breaches = [i for i in incidents if i.is_data_breach]
    data_breaches_pending = [i for i in data_breaches if not i.notification_date]
    data_breaches_notified = [i for i in data_breaches if i.notification_date]
    data_breaches_72h_violation = len([i for i in data_breaches
                                        if i.notification_date and
                                        (i.notification_date - i.discovery_date).total_seconds() > 72*3600])

    # Top incidentes críticos
    top_incidents = Incident.query.filter(
        Incident.severity.in_([IncidentSeverity.CRITICAL, IncidentSeverity.HIGH])
    ).order_by(Incident.severity.desc(), Incident.reported_date.desc()).limit(10).all()

    # Preparar datos para gráficos
    # Tendencia temporal
    trend_data = defaultdict(int)
    for incident in incidents:
        date_key = incident.reported_date.strftime('%Y-%m-%d')
        trend_data[date_key] += 1

    # Distribución por severidad
    severity_data = defaultdict(int)
    for incident in incidents:
        severity_data[incident.severity.value] += 1

    # Distribución por categoría
    category_data = defaultdict(int)
    for incident in incidents:
        category_data[incident.category.value] += 1

    # Distribución por estado
    status_data = defaultdict(int)
    for incident in incidents:
        status_data[incident.status.value] += 1

    # Impacto CIA
    cia_data = {
        'confidentiality': len([i for i in incidents if i.impact_confidentiality]),
        'integrity': len([i for i in incidents if i.impact_integrity]),
        'availability': len([i for i in incidents if i.impact_availability])
    }

    # Compliance ISO 27001
    compliance_5_24 = (len([i for i in incidents if i.description]) / total * 100) if total > 0 else 0
    compliance_5_25 = (len([i for i in incidents if i.status != IncidentStatus.NEW]) / total * 100) if total > 0 else 0
    compliance_5_27 = (len([i for i in incidents if i.lessons_learned]) / total * 100) if total > 0 else 0
    compliance_5_28 = (len([i for i in incidents if i.evidences]) / total * 100) if total > 0 else 0

    metrics = {
        'total': total,
        'critical': critical,
        'resolved': resolved,
        'resolution_rate': resolution_rate,
        'avg_response_time': avg_response_time,
        'avg_containment_time': avg_containment_time,
        'avg_resolution_time': avg_resolution_time,
        'avg_closure_time': avg_closure_time,
        'trend_total': trend_total,
        'data_breaches': len(data_breaches),
        'data_breaches_pending': len(data_breaches_pending),
        'data_breaches_notified': len(data_breaches_notified),
        'data_breaches_72h_violation': data_breaches_72h_violation,
        'compliance_5_24': compliance_5_24,
        'compliance_5_25': compliance_5_25,
        'compliance_5_27': compliance_5_27,
        'compliance_5_28': compliance_5_28
    }

    chart_data = {
        'trend': {
            'labels': sorted(trend_data.keys()),
            'data': [trend_data[k] for k in sorted(trend_data.keys())]
        },
        'severity': {
            'labels': list(severity_data.keys()),
            'data': list(severity_data.values())
        },
        'category': {
            'labels': list(category_data.keys()),
            'data': list(category_data.values())
        },
        'status': {
            'labels': list(status_data.keys()),
            'data': list(status_data.values())
        },
        'cia': {
            'data': [cia_data['confidentiality'], cia_data['integrity'], cia_data['availability']]
        }
    }

    return render_template('incidents/reports.html',
                          metrics=metrics,
                          chart_data=chart_data,
                          top_incidents=top_incidents,
                          data_breaches=data_breaches[:5],
                          period=period,
                          category=category,
                          start_date=start_date.strftime('%Y-%m-%d'),
                          end_date=end_date.strftime('%Y-%m-%d'),
                          now=datetime.utcnow(),
                          IncidentCategory=IncidentCategory)


# ============================================================================
# API ENDPOINTS
# ============================================================================

@incidents_bp.route('/api/stats')
@login_required
def api_stats():
    """API para obtener estadísticas"""
    stats = {
        'total': Incident.query.count(),
        'new': Incident.query.filter_by(status=IncidentStatus.NEW).count(),
        'in_progress': Incident.query.filter_by(status=IncidentStatus.IN_PROGRESS).count(),
        'resolved': Incident.query.filter_by(status=IncidentStatus.RESOLVED).count(),
        'closed': Incident.query.filter_by(status=IncidentStatus.CLOSED).count(),
        'critical': Incident.query.filter_by(severity=IncidentSeverity.CRITICAL).count()
    }
    return jsonify(stats)


@incidents_bp.route('/api/<int:id>')
@login_required
def api_get(id):
    """API para obtener un incidente"""
    incident = Incident.query.get_or_404(id)
    return jsonify(incident.to_dict())


@incidents_bp.route('/reports/export')
@login_required
def export_reports():
    """Exportar reportes en formato PDF o Excel"""
    from flask import make_response
    import io

    export_format = request.args.get('format', 'pdf')
    period = request.args.get('period', 30, type=int)
    category = request.args.get('category', 'all')

    # Calcular rango de fechas
    end_date = datetime.utcnow()
    if period == 'custom':
        start_date = datetime.strptime(request.args.get('start_date'), '%Y-%m-%d')
        end_date = datetime.strptime(request.args.get('end_date'), '%Y-%m-%d')
    else:
        start_date = end_date - timedelta(days=period)

    # Query base
    query = Incident.query.filter(Incident.reported_date >= start_date)
    if category != 'all':
        query = query.filter(Incident.category == IncidentCategory[category])

    incidents = query.all()

    if export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte de Incidentes"

            # Encabezados
            headers = ['Número', 'Título', 'Categoría', 'Severidad', 'Estado', 'Fecha Reporte',
                      'Fecha Descubrimiento', 'Responsable', 'Tiempo Resolución (h)']

            # Estilo de encabezados
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Datos
            for row, incident in enumerate(incidents, start=2):
                ws.cell(row=row, column=1, value=incident.incident_number)
                ws.cell(row=row, column=2, value=incident.title)
                ws.cell(row=row, column=3, value=incident.category.value if incident.category else '')
                ws.cell(row=row, column=4, value=incident.severity.value if incident.severity else '')
                ws.cell(row=row, column=5, value=incident.status.value if incident.status else '')
                ws.cell(row=row, column=6, value=incident.reported_date.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=7, value=incident.discovery_date.strftime('%Y-%m-%d') if incident.discovery_date else '')
                ws.cell(row=row, column=8, value=incident.assigned_to.name if incident.assigned_to else 'Sin asignar')
                ws.cell(row=row, column=9, value=incident.calculate_resolution_time() or '')

            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Guardar en memoria
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = make_response(output.read())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename=reporte_incidentes_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'

            return response

        except ImportError:
            flash('openpyxl no está instalado. Instale con: pip install openpyxl', 'error')
            return redirect(url_for('incidents.reports'))

    elif export_format == 'pdf':
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch

            # Crear PDF en memoria
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []

            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#4472C4'),
                spaceAfter=30,
            )

            # Título
            title = Paragraph(f"Reporte de Incidentes de Seguridad", title_style)
            elements.append(title)

            subtitle = Paragraph(
                f"Periodo: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}<br/>"
                f"Total de incidentes: {len(incidents)}",
                styles['Normal']
            )
            elements.append(subtitle)
            elements.append(Spacer(1, 20))

            # Tabla de datos
            data = [['Número', 'Título', 'Severidad', 'Estado', 'Fecha']]

            for incident in incidents:
                data.append([
                    incident.incident_number,
                    incident.title[:40] + '...' if len(incident.title) > 40 else incident.title,
                    incident.severity.value if incident.severity else '',
                    incident.status.value if incident.status else '',
                    incident.reported_date.strftime('%d/%m/%Y')
                ])

            table = Table(data, colWidths=[1.2*inch, 3*inch, 1*inch, 1.2*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(table)

            # Generar PDF
            doc.build(elements)
            buffer.seek(0)

            response = make_response(buffer.read())
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename=reporte_incidentes_{datetime.utcnow().strftime("%Y%m%d")}.pdf'

            return response

        except ImportError:
            flash('reportlab no está instalado. Instale con: pip install reportlab', 'error')
            return redirect(url_for('incidents.reports'))

    else:
        flash('Formato de exportación no soportado', 'error')
        return redirect(url_for('incidents.reports'))
